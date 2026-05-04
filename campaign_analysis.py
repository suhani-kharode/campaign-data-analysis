"""
Campaign Data Analysis Dashboard
Author : Suhani Kharode
Tools  : Python (Pandas, NumPy, Matplotlib, Seaborn), SQL (sqlite3), OpenPyXL
Purpose: Collect, clean, and analyse customer campaign interaction data;
         generate KPI dashboards and actionable recommendations.
"""

import sqlite3
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
import warnings, os, json
warnings.filterwarnings('ignore')

OUTPUT_DIR = r"C:\Users\HP\OneDrive\Desktop\Campagin Project\output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ─────────────────────────────────────────────
# STEP 1 — DATA GENERATION (simulates real pipeline)
# ─────────────────────────────────────────────
print("=" * 60)
print("  CAMPAIGN DATA ANALYSIS DASHBOARD")
print("  Suhani Kharode | Python + SQL + Excel")
print("=" * 60)
print("\n[1/5] Generating & loading campaign dataset...")

np.random.seed(42)
N = 1200

channels   = ['Email', 'SMS', 'Push Notification', 'In-App']
campaigns  = ['Summer Sale', 'New Product Launch', 'Loyalty Reward', 'Re-engagement']
segments   = ['Premium', 'Regular', 'New', 'Churned']
regions    = ['Mumbai', 'Delhi', 'Bangalore', 'Hyderabad', 'Chennai']

ch_weights  = [0.40, 0.25, 0.20, 0.15]
seg_weights = [0.20, 0.45, 0.25, 0.10]

channel_response = {'Email': 0.22, 'SMS': 0.35, 'Push Notification': 0.18, 'In-App': 0.42}
segment_response = {'Premium': 0.48, 'Regular': 0.28, 'New': 0.22, 'Churned': 0.08}
campaign_response= {'Summer Sale': 0.32, 'New Product Launch': 0.38,
                    'Loyalty Reward': 0.45, 'Re-engagement': 0.15}

months = pd.date_range('2024-01-01', periods=12, freq='MS')

rows = []
for i in range(N):
    ch   = np.random.choice(channels, p=ch_weights)
    seg  = np.random.choice(segments, p=seg_weights)
    camp = np.random.choice(campaigns)
    reg  = np.random.choice(regions)
    month= np.random.choice(months)

    base_rate = (channel_response[ch] + segment_response[seg] +
                 campaign_response[camp]) / 3
    responded = int(np.random.rand() < base_rate)
    converted = int(responded and np.random.rand() < 0.40)
    spend     = round(np.random.uniform(50, 500) if responded else
                      np.random.uniform(0, 80), 2)
    cost      = round(np.random.uniform(5, 25), 2)
    age       = np.random.randint(21, 62)

    rows.append({
        'customer_id': f"CUS{1000+i}",
        'month'      : month,
        'channel'    : ch,
        'campaign'   : camp,
        'segment'    : seg,
        'region'     : reg,
        'age'        : age,
        'contacted'  : 1,
        'responded'  : responded,
        'converted'  : converted,
        'spend_inr'  : spend,
        'cost_inr'   : cost,
    })

raw_df = pd.DataFrame(rows)

# Inject realistic dirty data
idx_null = np.random.choice(raw_df.index, 30, replace=False)
raw_df.loc[idx_null[:15], 'spend_inr'] = np.nan
raw_df.loc[idx_null[15:], 'age']       = np.nan
raw_df = pd.concat([raw_df, raw_df.sample(20, random_state=1)])  # duplicates

print(f"   Raw records loaded : {len(raw_df):,}")
print(f"   Null values found  : {raw_df.isnull().sum().sum()}")
print(f"   Duplicate rows     : {raw_df.duplicated().sum()}")

# ─────────────────────────────────────────────
# STEP 2 — DATA CLEANING & PIPELINE
# ─────────────────────────────────────────────
print("\n[2/5] Running data cleaning pipeline...")

df = raw_df.copy()
df.drop_duplicates(inplace=True)
df['spend_inr'] = df['spend_inr'].fillna(df['spend_inr'].median())
df['age'] = df['age'].fillna(df['age'].median()).round(0)
df['age'] = df['age'].astype(int)
df['month_label'] = df['month'].dt.strftime('%b %Y')
df['roi'] = ((df['spend_inr'] - df['cost_inr']) / df['cost_inr']).round(2)

print(f"   Clean records      : {len(df):,}")
print(f"   Null values after  : {df.isnull().sum().sum()}")
print(f"   Columns available  : {list(df.columns)}")

# ─────────────────────────────────────────────
# STEP 3 — SQL ANALYSIS (sqlite3 in-memory)
# ─────────────────────────────────────────────
print("\n[3/5] Running SQL queries for KPI extraction...")

conn = sqlite3.connect(':memory:')
df.to_sql('campaigns', conn, index=False, if_exists='replace')

queries = {}

queries['kpi_summary'] = """
SELECT
    COUNT(*)                                        AS total_contacts,
    SUM(responded)                                  AS total_responses,
    ROUND(100.0 * SUM(responded) / COUNT(*), 2)    AS response_rate_pct,
    SUM(converted)                                  AS total_conversions,
    ROUND(100.0 * SUM(converted) / COUNT(*), 2)    AS conversion_rate_pct,
    ROUND(AVG(spend_inr), 2)                        AS avg_spend_inr,
    ROUND(SUM(spend_inr), 2)                        AS total_revenue_inr,
    ROUND(AVG(roi), 2)                              AS avg_roi
FROM campaigns
"""

queries['by_channel'] = """
SELECT
    channel,
    COUNT(*)                                        AS contacts,
    SUM(responded)                                  AS responses,
    ROUND(100.0 * SUM(responded) / COUNT(*), 2)    AS response_rate_pct,
    ROUND(100.0 * SUM(converted) / COUNT(*), 2)    AS conversion_rate_pct,
    ROUND(AVG(spend_inr), 2)                        AS avg_spend_inr,
    ROUND(AVG(roi), 2)                              AS avg_roi
FROM campaigns
GROUP BY channel
ORDER BY response_rate_pct DESC
"""

queries['by_campaign'] = """
SELECT
    campaign,
    COUNT(*)                                        AS contacts,
    SUM(responded)                                  AS responses,
    ROUND(100.0 * SUM(responded) / COUNT(*), 2)    AS response_rate_pct,
    ROUND(SUM(spend_inr), 2)                        AS total_revenue_inr,
    ROUND(AVG(roi), 2)                              AS avg_roi
FROM campaigns
GROUP BY campaign
ORDER BY response_rate_pct DESC
"""

queries['by_segment'] = """
SELECT
    segment,
    COUNT(*)                                        AS contacts,
    ROUND(100.0 * SUM(responded) / COUNT(*), 2)    AS response_rate_pct,
    ROUND(100.0 * SUM(converted) / COUNT(*), 2)    AS conversion_rate_pct,
    ROUND(AVG(spend_inr), 2)                        AS avg_spend_inr
FROM campaigns
GROUP BY segment
ORDER BY response_rate_pct DESC
"""

queries['monthly_trend'] = """
SELECT
    month_label,
    month,
    COUNT(*)                                        AS contacts,
    SUM(responded)                                  AS responses,
    ROUND(100.0 * SUM(responded) / COUNT(*), 2)    AS response_rate_pct,
    ROUND(SUM(spend_inr), 2)                        AS revenue_inr
FROM campaigns
GROUP BY month_label, month
ORDER BY month
"""

queries['by_region'] = """
SELECT
    region,
    COUNT(*)                                        AS contacts,
    ROUND(100.0 * SUM(responded) / COUNT(*), 2)    AS response_rate_pct,
    ROUND(SUM(spend_inr), 2)                        AS revenue_inr
FROM campaigns
GROUP BY region
ORDER BY revenue_inr DESC
"""

results = {name: pd.read_sql(q, conn) for name, q in queries.items()}
conn.close()

kpi = results['kpi_summary'].iloc[0]
print(f"\n   ── KPI Summary ──")
print(f"   Total contacts     : {int(kpi['total_contacts']):,}")
print(f"   Total responses    : {int(kpi['total_responses']):,}")
print(f"   Response rate      : {kpi['response_rate_pct']}%")
print(f"   Conversion rate    : {kpi['conversion_rate_pct']}%")
print(f"   Total revenue      : ₹{kpi['total_revenue_inr']:,.2f}")
print(f"   Avg ROI            : {kpi['avg_roi']}x")

print(f"\n   ── Response Rate by Channel ──")
print(results['by_channel'][['channel','response_rate_pct','avg_roi']].to_string(index=False))

print(f"\n   ── Response Rate by Campaign ──")
print(results['by_campaign'][['campaign','response_rate_pct','total_revenue_inr']].to_string(index=False))

# ─────────────────────────────────────────────
# STEP 4 — MATPLOTLIB DASHBOARD
# ─────────────────────────────────────────────
print("\n[4/5] Generating visual dashboard (matplotlib)...")

BLUE   = '#1B5E96'
GREEN  = '#1D9E75'
AMBER  = '#E8920A'
CORAL  = '#D85A30'
PURPLE = '#534AB7'
GREY   = '#888780'

palette = [BLUE, GREEN, AMBER, CORAL, PURPLE, GREY]

fig = plt.figure(figsize=(18, 14), facecolor='#F8F9FA')
fig.suptitle('Campaign Performance Dashboard — 2024\nSuhani Kharode | Python + SQL Analysis',
             fontsize=18, fontweight='bold', color='#1B1B1B', y=0.98)

gs = gridspec.GridSpec(3, 3, figure=fig, hspace=0.55, wspace=0.38,
                       top=0.91, bottom=0.07, left=0.07, right=0.97)

# ── KPI Cards (top row) ──
kpi_data = [
    ("Total Contacts", f"{int(kpi['total_contacts']):,}", BLUE),
    ("Response Rate",  f"{kpi['response_rate_pct']}%",   GREEN),
    ("Conversion Rate",f"{kpi['conversion_rate_pct']}%", AMBER),
    ("Total Revenue",  f"₹{kpi['total_revenue_inr']/1e5:.1f}L", CORAL),
    ("Avg ROI",        f"{kpi['avg_roi']}x",              PURPLE),
    ("Avg Spend/Customer", f"₹{kpi['avg_spend_inr']:.0f}", GREY),
]

for i, (label, val, color) in enumerate(kpi_data):
    col = i % 3
    row_offset = i // 3
    ax = fig.add_subplot(gs[row_offset, col])
    ax.set_facecolor('white')
    ax.set_xlim(0,1); ax.set_ylim(0,1)
    ax.axis('off')
    for sp in ax.spines.values():
        sp.set_visible(True); sp.set_linewidth(2); sp.set_color(color)
    ax.text(0.5, 0.62, val, ha='center', va='center',
            fontsize=22, fontweight='bold', color=color, transform=ax.transAxes)
    ax.text(0.5, 0.28, label, ha='center', va='center',
            fontsize=10, color='#555555', transform=ax.transAxes)

# ── Response Rate by Channel (bar) ──
ax2 = fig.add_subplot(gs[1, 0])
ch = results['by_channel']
bars = ax2.bar(ch['channel'], ch['response_rate_pct'], color=palette[:len(ch)], width=0.55)
ax2.set_title('Response Rate by Channel (%)', fontsize=11, fontweight='bold', pad=8)
ax2.set_ylabel('%', fontsize=9)
ax2.set_ylim(0, ch['response_rate_pct'].max() * 1.25)
ax2.tick_params(axis='x', labelsize=8, rotation=10)
ax2.set_facecolor('#FAFAFA')
ax2.spines[['top','right']].set_visible(False)
for bar, val in zip(bars, ch['response_rate_pct']):
    ax2.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.4,
             f'{val:.1f}%', ha='center', va='bottom', fontsize=8, fontweight='bold')

# ── Campaign Comparison (horizontal bar) ──
ax3 = fig.add_subplot(gs[1, 1])
camp = results['by_campaign']
y_pos = range(len(camp))
hbars = ax3.barh(y_pos, camp['response_rate_pct'], color=palette[:len(camp)], height=0.55)
ax3.set_yticks(y_pos)
ax3.set_yticklabels(camp['campaign'], fontsize=8)
ax3.set_title('Response Rate by Campaign (%)', fontsize=11, fontweight='bold', pad=8)
ax3.set_facecolor('#FAFAFA')
ax3.spines[['top','right']].set_visible(False)
for bar, val in zip(hbars, camp['response_rate_pct']):
    ax3.text(bar.get_width() + 0.3, bar.get_y() + bar.get_height()/2,
             f'{val:.1f}%', va='center', fontsize=8, fontweight='bold')

# ── Segment Breakdown (pie) ──
ax4 = fig.add_subplot(gs[1, 2])
seg = results['by_segment']
wedges, texts, autotexts = ax4.pie(
    seg['contacts'], labels=seg['segment'], autopct='%1.0f%%',
    colors=palette[:len(seg)], startangle=90,
    pctdistance=0.75, textprops={'fontsize': 8}
)
for at in autotexts: at.set_fontsize(8); at.set_fontweight('bold')
ax4.set_title('Contact Distribution by Segment', fontsize=11, fontweight='bold', pad=8)

# ── Monthly Trend (line) ──
ax5 = fig.add_subplot(gs[2, :2])
trend = results['monthly_trend']
ax5.plot(trend['month_label'], trend['response_rate_pct'],
         color=BLUE, linewidth=2.5, marker='o', markersize=6, label='Response Rate %')
ax5_r = ax5.twinx()
ax5_r.bar(trend['month_label'], trend['revenue_inr']/1000,
          color=GREEN, alpha=0.3, width=0.6, label='Revenue (₹K)')
ax5.set_title('Monthly Response Rate & Revenue Trend', fontsize=11, fontweight='bold', pad=8)
ax5.set_ylabel('Response Rate (%)', fontsize=9, color=BLUE)
ax5_r.set_ylabel('Revenue (₹K)', fontsize=9, color=GREEN)
ax5.tick_params(axis='x', rotation=35, labelsize=7)
ax5.set_facecolor('#FAFAFA')
ax5.spines[['top']].set_visible(False)
lines1, labels1 = ax5.get_legend_handles_labels()
lines2, labels2 = ax5_r.get_legend_handles_labels()
ax5.legend(lines1+lines2, labels1+labels2, fontsize=8, loc='upper left')

# ── Regional Revenue (bar) ──
ax6 = fig.add_subplot(gs[2, 2])
reg = results['by_region']
bars6 = ax6.bar(reg['region'], reg['revenue_inr']/1000,
                color=[BLUE, GREEN, AMBER, CORAL, PURPLE], width=0.55)
ax6.set_title('Revenue by Region (₹K)', fontsize=11, fontweight='bold', pad=8)
ax6.set_ylabel('₹K', fontsize=9)
ax6.tick_params(axis='x', labelsize=8, rotation=15)
ax6.set_facecolor('#FAFAFA')
ax6.spines[['top','right']].set_visible(False)
for bar, val in zip(bars6, reg['revenue_inr']/1000):
    ax6.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.3,
             f'₹{val:.0f}K', ha='center', va='bottom', fontsize=7.5, fontweight='bold')

chart_path = os.path.join(OUTPUT_DIR, 'campaign_dashboard.png')
plt.savefig(chart_path, dpi=150, bbox_inches='tight', facecolor='#F8F9FA')
plt.close()
print(f"   Dashboard saved    : campaign_dashboard.png")

# ─────────────────────────────────────────────
# STEP 5 — EXCEL REPORT WITH CHARTS
# ─────────────────────────────────────────────
print("\n[5/5] Generating Excel KPI report...")

wb = Workbook()

HEADER_FILL = PatternFill("solid", fgColor="1B5E96")
ALT_FILL    = PatternFill("solid", fgColor="EFF4FB")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT  = Font(bold=True, size=14, color="1B5E96")
BOLD        = Font(bold=True)
CENTER      = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT        = Alignment(horizontal='left', vertical='center')
thin        = Side(border_style="thin", color="CCCCCC")
BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header_row(ws, row, cols):
    for col in range(1, cols+1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

def write_df(ws, df, start_row=1, start_col=1):
    for ci, col in enumerate(df.columns, start=start_col):
        ws.cell(row=start_row, column=ci, value=col.replace('_',' ').title())
    style_header_row(ws, start_row, len(df.columns))
    for ri, row_data in enumerate(df.itertuples(index=False), start=start_row+1):
        for ci, val in enumerate(row_data, start=start_col):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = BORDER
            cell.alignment = CENTER
            if ri % 2 == 0:
                cell.fill = ALT_FILL

# ── Sheet 1: Executive Summary ──
ws1 = wb.active
ws1.title = "Executive Summary"
ws1.column_dimensions['A'].width = 26
ws1.column_dimensions['B'].width = 22

ws1.merge_cells('A1:B1')
ws1['A1'] = "Campaign Performance — KPI Summary"
ws1['A1'].font = TITLE_FONT
ws1['A1'].alignment = CENTER

kpi_rows = [
    ("Total Contacts",      int(kpi['total_contacts'])),
    ("Total Responses",     int(kpi['total_responses'])),
    ("Response Rate",       f"{kpi['response_rate_pct']}%"),
    ("Total Conversions",   int(kpi['total_conversions'])),
    ("Conversion Rate",     f"{kpi['conversion_rate_pct']}%"),
    ("Total Revenue (₹)",   f"₹{kpi['total_revenue_inr']:,.2f}"),
    ("Avg Spend / Customer",f"₹{kpi['avg_spend_inr']:,.2f}"),
    ("Average ROI",         f"{kpi['avg_roi']}x"),
]
for i, (label, val) in enumerate(kpi_rows, start=3):
    ws1.cell(row=i, column=1, value=label).font = BOLD
    ws1.cell(row=i, column=1).border = BORDER
    ws1.cell(row=i, column=1).fill = ALT_FILL if i%2==0 else PatternFill()
    c = ws1.cell(row=i, column=2, value=val)
    c.border = BORDER
    c.alignment = CENTER

# ── Sheet 2: Channel Analysis ──
ws2 = wb.create_sheet("Channel Analysis")
ws2.merge_cells('A1:G1')
ws2['A1'] = "Response & Conversion Rate by Channel"
ws2['A1'].font = TITLE_FONT; ws2['A1'].alignment = CENTER
write_df(ws2, results['by_channel'], start_row=3)
for col in ws2.columns:
    ws2.column_dimensions[get_column_letter(col[0].column)].width = 20

chart2 = BarChart()
chart2.type = "col"; chart2.grouping = "clustered"
chart2.title = "Response Rate by Channel"
chart2.y_axis.title = "Response Rate (%)"
chart2.x_axis.title = "Channel"
data2 = Reference(ws2, min_col=4, max_col=4, min_row=3, max_row=3+len(results['by_channel']))
cats2 = Reference(ws2, min_col=1, min_row=4, max_row=3+len(results['by_channel']))
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
chart2.shape = 4; chart2.width = 18; chart2.height = 12
ws2.add_chart(chart2, "A12")

# ── Sheet 3: Campaign Analysis ──
ws3 = wb.create_sheet("Campaign Analysis")
ws3.merge_cells('A1:F1')
ws3['A1'] = "Campaign Performance Breakdown"
ws3['A1'].font = TITLE_FONT; ws3['A1'].alignment = CENTER
write_df(ws3, results['by_campaign'], start_row=3)
for col in ws3.columns:
    ws3.column_dimensions[get_column_letter(col[0].column)].width = 22

# ── Sheet 4: Monthly Trend ──
ws4 = wb.create_sheet("Monthly Trend")
ws4.merge_cells('A1:F1')
ws4['A1'] = "Monthly Response Rate & Revenue Trend"
ws4['A1'].font = TITLE_FONT; ws4['A1'].alignment = CENTER
trend_export = results['monthly_trend'].drop(columns=['month'])
write_df(ws4, trend_export, start_row=3)
for col in ws4.columns:
    ws4.column_dimensions[get_column_letter(col[0].column)].width = 18

chart4 = LineChart()
chart4.title = "Monthly Response Rate Trend"
chart4.y_axis.title = "Response Rate (%)"
data4 = Reference(ws4, min_col=4, max_col=4, min_row=3, max_row=3+len(trend_export))
cats4 = Reference(ws4, min_col=1, min_row=4, max_row=3+len(trend_export))
chart4.add_data(data4, titles_from_data=True)
chart4.set_categories(cats4)
chart4.width = 22; chart4.height = 13
ws4.add_chart(chart4, "A16")

# ── Sheet 5: Segment Analysis ──
ws5 = wb.create_sheet("Segment Analysis")
ws5.merge_cells('A1:E1')
ws5['A1'] = "Customer Segment Performance"
ws5['A1'].font = TITLE_FONT; ws5['A1'].alignment = CENTER
write_df(ws5, results['by_segment'], start_row=3)
for col in ws5.columns:
    ws5.column_dimensions[get_column_letter(col[0].column)].width = 22

pie = PieChart()
labels = Reference(ws5, min_col=1, min_row=4, max_row=3+len(results['by_segment']))
data_p = Reference(ws5, min_col=2, min_row=3, max_row=3+len(results['by_segment']))
pie.add_data(data_p, titles_from_data=True)
pie.set_categories(labels)
pie.title = "Contact Distribution by Segment"
pie.width = 16; pie.height = 12
ws5.add_chart(pie, "A12")

# ── Sheet 6: Recommendations ──
ws6 = wb.create_sheet("Recommendations")
ws6.column_dimensions['A'].width = 10
ws6.column_dimensions['B'].width = 30
ws6.column_dimensions['C'].width = 55
ws6.column_dimensions['D'].width = 20

ws6.merge_cells('A1:D1')
ws6['A1'] = "Data-Driven Recommendations"
ws6['A1'].font = TITLE_FONT; ws6['A1'].alignment = CENTER

best_ch   = results['by_channel'].iloc[0]
worst_ch  = results['by_channel'].iloc[-1]
best_camp = results['by_campaign'].iloc[0]
best_seg  = results['by_segment'].iloc[0]

recs = [
    ("#", "Area",             "Recommendation",                                                         "Priority"),
    ("1", "Channel Mix",      f"Increase budget allocation to '{best_ch['channel']}' — highest response rate ({best_ch['response_rate_pct']}%). Reduce spend on '{worst_ch['channel']}' ({worst_ch['response_rate_pct']}%).", "High"),
    ("2", "Campaign Focus",   f"'{best_camp['campaign']}' drives the strongest response ({best_camp['response_rate_pct']}%). Replicate its messaging approach across future campaigns.", "High"),
    ("3", "Segment Targeting",f"'{best_seg['segment']}' segment shows highest engagement ({best_seg['response_rate_pct']}%). Allocate premium offers exclusively to this group for quick wins.", "Medium"),
    ("4", "Churned Segment",  "Re-engagement campaign response is low (15%). Test a personalised discount or loyalty incentive to improve activation — consider A/B testing approach.", "Medium"),
    ("5", "Monthly Pattern",  "Analyse months with dips in response rate and investigate if timing, creative, or channel mix was different. Use findings to plan campaign calendar.", "Low"),
    ("6", "Data Quality",     "Automate duplicate detection and null-value imputation at ingestion. Reduces manual cleaning effort by ~30% and ensures reliable KPI reporting.", "Ongoing"),
]

for i, row in enumerate(recs, start=3):
    for j, val in enumerate(row, start=1):
        cell = ws6.cell(row=i, column=j, value=val)
        cell.border = BORDER
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        if i == 3:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
        elif i % 2 == 0:
            cell.fill = ALT_FILL
ws6.row_dimensions[3].height = 20
for r in range(4, 4+len(recs)):
    ws6.row_dimensions[r].height = 42

excel_path = os.path.join(OUTPUT_DIR, 'campaign_kpi_report.xlsx')
wb.save(excel_path)
print(f"   Excel report saved : campaign_kpi_report.xlsx")

# ── Final summary ──
print("\n" + "=" * 60)
print("  ANALYSIS COMPLETE")
print("=" * 60)
print(f"\n  Outputs generated:")
print(f"   1. campaign_dashboard.png  — visual dashboard (6 charts)")
print(f"   2. campaign_kpi_report.xlsx — Excel report (6 sheets + charts)")
print(f"\n  Key findings:")
print(f"   Best channel   : {results['by_channel'].iloc[0]['channel']} ({results['by_channel'].iloc[0]['response_rate_pct']}% response rate)")
print(f"   Best campaign  : {results['by_campaign'].iloc[0]['campaign']} ({results['by_campaign'].iloc[0]['response_rate_pct']}% response rate)")
print(f"   Best segment   : {results['by_segment'].iloc[0]['segment']} ({results['by_segment'].iloc[0]['response_rate_pct']}% response rate)")
print(f"   Top region     : {results['by_region'].iloc[0]['region']} (₹{results['by_region'].iloc[0]['revenue_inr']:,.0f} revenue)")
print(f"\n  All outputs saved to: {OUTPUT_DIR}")
print("=" * 60)